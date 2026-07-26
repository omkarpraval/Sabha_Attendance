import io
import base64
from typing import List
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image as RLImage, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

def generate_excel_report(events_grouped: List[dict]) -> bytes:
    """
    Generates a formatted Excel spreadsheet grouped by Event with top header details per event.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sabha Attendance"

    # Fonts & Styles
    title_font = Font(name="Arial", size=14, bold=True, color="8B3A3A")
    event_title_font = Font(name="Arial", size=12, bold=True, color="8B3A3A")
    subtitle_font = Font(name="Arial", size=10, bold=True, color="3A322C")
    meta_font = Font(name="Arial", size=10, italic=True, color="555555")
    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="8B3A3A", end_color="8B3A3A", fill_type="solid")
    
    fill_present = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")
    fill_absent = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    fill_excused = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    
    border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )

    current_row = 1

    # Main Document Banner
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    cell = ws.cell(row=1, column=1, value="BAPS SABHA ATTENDANCE OFFICIAL REPORT")
    cell.font = title_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    current_row = 3

    for ev in events_grouped:
        # Event Header Box
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
        c_title = ws.cell(row=current_row, column=1, value=f"EVENT: {ev['event_title']}")
        c_title.font = event_title_font
        c_title.fill = PatternFill(start_color="FDFBF7", end_color="FDFBF7", fill_type="solid")
        current_row += 1

        # Event Meta Row (Location & Date)
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
        meta_text = f"Date: {ev['event_date']} ({ev['start_time']} - {ev['end_time']} IST)   |   Location / Venue: {ev['venue_name']}"
        c_meta = ws.cell(row=current_row, column=1, value=meta_text)
        c_meta.font = meta_font
        current_row += 1

        # Turnout Metrics Row
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
        summary_text = f"Total Headcount: {ev['total_headcount']}   |   Present: {ev['present_count']}   |   Absent: {ev['absent_count']}   |   Turnout: {ev['turnout_pct']}%"
        c_sum = ws.cell(row=current_row, column=1, value=summary_text)
        c_sum.font = subtitle_font
        current_row += 1

        # Table Column Headers
        headers = ["Member Name", "Phone Number", "Status", "Marked By", "Method", "Distance (m)", "Timestamp (UTC)"]
        for col_idx, h in enumerate(headers, 1):
            c_h = ws.cell(row=current_row, column=col_idx, value=h)
            c_h.font = header_font
            c_h.fill = header_fill
            c_h.alignment = Alignment(horizontal="center", vertical="center")
        current_row += 1

        # Attendance Table Rows for this event
        if not ev["records"]:
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
            c_empty = ws.cell(row=current_row, column=1, value="No attendance records logged for this event.")
            c_empty.font = Font(name="Arial", size=10, italic=True)
            current_row += 1
        else:
            for r in ev["records"]:
                status_str = r.get("status", "").upper()
                ws.append([
                    r.get("user_name", ""),
                    r.get("user_phone", ""),
                    status_str,
                    r.get("marked_by_name", ""),
                    r.get("marking_method", "").replace("_", " ").title(),
                    r.get("distance_meters", "N/A"),
                    r.get("timestamp_utc", "")
                ])
                
                for col_idx in range(1, 8):
                    c = ws.cell(row=current_row, column=col_idx)
                    c.border = border
                    c.alignment = Alignment(vertical="center")
                
                status_cell = ws.cell(row=current_row, column=3)
                if status_str == "PRESENT":
                    status_cell.fill = fill_present
                elif status_str == "ABSENT":
                    status_cell.fill = fill_absent
                elif status_str == "EXCUSED":
                    status_cell.fill = fill_excused
                
                current_row += 1

        # Blank rows separator between events
        current_row += 2

    # Auto-fit column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 16)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def generate_pdf_report(events_grouped: List[dict], title_suffix: str = "") -> bytes:
    """
    Generates an event-grouped PDF report with Event Name, Location, and Date at the top of each event.
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    story = []

    styles = getSampleStyleSheet()
    doc_title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=18,
        textColor=colors.HexColor('#8B3A3A'),
        spaceAfter=4
    )
    doc_sub_style = ParagraphStyle(
        'DocSub',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=10,
        textColor=colors.HexColor('#3A322C'),
        spaceAfter=15
    )

    ev_title_style = ParagraphStyle(
        'EvTitle',
        parent=styles['Heading2'],
        fontName='Helvetica-Bold',
        fontSize=13,
        textColor=colors.HexColor('#8B3A3A'),
        spaceAfter=3
    )

    ev_meta_style = ParagraphStyle(
        'EvMeta',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9,
        textColor=colors.HexColor('#3A322C'),
        spaceAfter=4
    )

    ev_metrics_style = ParagraphStyle(
        'EvMetrics',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        textColor=colors.HexColor('#5B8C5B'),
        spaceAfter=8
    )

    story.append(Paragraph("Sabha Attendance Official Summary Report", doc_title_style))
    story.append(Paragraph(f"Report Summary {title_suffix} | Total Events Included: {len(events_grouped)}", doc_sub_style))
    story.append(Spacer(1, 5))

    for idx, ev in enumerate(events_grouped):
        if idx > 0:
            story.append(Spacer(1, 15))

        # Event Name, Location, Date at the top of each event
        story.append(Paragraph(f"Event: {ev['event_title']}", ev_title_style))
        story.append(Paragraph(f"Date: {ev['event_date']} ({ev['start_time']} - {ev['end_time']} IST)   |   Location / Venue: {ev['venue_name']}", ev_meta_style))
        story.append(Paragraph(f"Total Headcount: <b>{ev['total_headcount']}</b>   |   Present: <b>{ev['present_count']}</b>   |   Absent: <b>{ev['absent_count']}</b>   (Turnout: {ev['turnout_pct']}%)", ev_metrics_style))

        # Table data
        table_data = [["Member Name", "Phone Number", "Status", "Marked By", "Method", "Distance (m)"]]
        
        if not ev["records"]:
            table_data.append(["No records logged for this event", "-", "-", "-", "-", "-"])
        else:
            for r in ev["records"]:
                table_data.append([
                    r.get("user_name", ""),
                    r.get("user_phone", ""),
                    r.get("status", "").upper(),
                    r.get("marked_by_name", ""),
                    r.get("marking_method", "").replace("_", " ").title(),
                    r.get("distance_meters", "N/A")
                ])

        t = Table(table_data, colWidths=[130, 90, 60, 110, 85, 60])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#8B3A3A')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 5),
            ('TOPPADDING', (0, 0), (-1, 0), 5),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E0E0E0')),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
        ]))
        story.append(t)

    doc.build(story)
    return buffer.getvalue()


def generate_qr_poster_pdf(
    event_title: str,
    event_date: str,
    start_time: str,
    end_time: str,
    venue_name: str,
    qr_ref: str,
    qr_base64_str: str
) -> bytes:
    """
    Generates a PDF Poster for a Sabha QR Code matching the exact portal layout and styling.
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=36,
        leftMargin=36,
        topMargin=36,
        bottomMargin=36
    )
    story = []
    styles = getSampleStyleSheet()

    header_title_style = ParagraphStyle(
        'HeaderTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=22,
        textColor=colors.HexColor('#8B3A3A'),
        alignment=1,
        spaceAfter=4
    )

    header_sub_style = ParagraphStyle(
        'HeaderSub',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=11,
        textColor=colors.HexColor('#E8A33D'),
        alignment=1,
        spaceAfter=15
    )

    event_title_style = ParagraphStyle(
        'EventTitle',
        parent=styles['Heading2'],
        fontName='Helvetica-Bold',
        fontSize=20,
        textColor=colors.HexColor('#3A322C'),
        alignment=1,
        spaceAfter=6
    )

    event_meta_style = ParagraphStyle(
        'EventMeta',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=11,
        textColor=colors.HexColor('#8B3A3A'),
        alignment=1,
        spaceAfter=15
    )

    ref_style = ParagraphStyle(
        'RefStyle',
        parent=styles['Code'],
        fontName='Courier-Bold',
        fontSize=12,
        textColor=colors.HexColor('#8B3A3A'),
        alignment=1
    )

    instruction_style = ParagraphStyle(
        'Instruction',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=11,
        textColor=colors.HexColor('#3A322C'),
        alignment=1,
        spaceAfter=6
    )

    story.append(Paragraph("BAPS SABHA ATTENDANCE SYSTEM", header_title_style))
    story.append(Paragraph("AUTOMATIC QR & GEOFENCE SYSTEM", header_sub_style))
    story.append(Spacer(1, 10))

    story.append(Paragraph(event_title, event_title_style))
    story.append(Paragraph(f"Date: {event_date} ({start_time} - {end_time} IST) | Venue: {venue_name}", event_meta_style))
    story.append(Spacer(1, 10))

    clean_b64 = qr_base64_str.split(",")[-1] if "," in qr_base64_str else qr_base64_str
    img_data = base64.b64decode(clean_b64)
    img_buffer = io.BytesIO(img_data)
    
    rl_img = RLImage(img_buffer, width=280, height=280)
    story.append(rl_img)
    story.append(Spacer(1, 20))

    story.append(Paragraph("Scan with Sabha App inside Mandir boundary to mark attendance", instruction_style))
    story.append(Spacer(1, 12))

    ref_table = Table([[Paragraph(f"REF: {qr_ref}", ref_style)]], colWidths=[360])
    ref_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#EFE7DA')),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#8B3A3A')),
    ]))
    story.append(ref_table)

    doc.build(story)
    return buffer.getvalue()
